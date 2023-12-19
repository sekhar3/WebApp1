import json
import openpyxl

"""
@author        : Indudhara
Description    : This class is used to read the test data 
"""


class ReadData:
    """@author : Sreenivas Reddy
    Description: This method is used to read test data from json file"""

    def get_test_data(self):
        f = open(r'..\..\Resources\Testdata.json')
        data = json.load(f)
        f.close()
        return data

    """@author : Sreenivas Reddy
    Description: This method is used to read test data from excel file
    """

    def get_excel_data(self, x, y):
        path = r"..\..\Resources\data.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        return sheet.cell(x, y).value

    """@author : Sreenivas Reddy
        Description: This method is used to read test data from excel file
    """

    def get_excel_data1(self, x, y):
        path = r"..\..\Resources\id.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        return sheet.cell(x, y).value

    """
        @author    : Indudhara
        Description   : This method is used to write data into excel file
        i/p parameter : row, column number and value to be saved in excel file
    """

    def write_data_into_excel(self, row, col, data):
        path = r"..\..\Resources\data.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet.cell(row=row, column=col).value = data
        wb.save(path)
        wb.close()

    def write_data_into_excel1(self, row, col, data):
        path = r"..\..\Resources\id.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet.cell(row=row, column=col).value = data
        wb.save(path)
        wb.close()